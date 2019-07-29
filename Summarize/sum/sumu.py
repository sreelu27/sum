from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize, sent_tokenize
import openpyxl


""" Creating a dictionary for storing the word frequency table."""
def generate_frequency_table(input_text) -> dict:
   
    """ Pre processing of data: Removing stop words"""
    stopWords = set(stopwords.words("english"))
    
    """Pre processing of data: Tokenization """
    words = word_tokenize(input_text)
    
    
    """ Pre processing of data: Stemming"""
    ps = PorterStemmer()
   

    frequencyTable = dict()
    for w in words:
        w = ps.stem(w)
        if w in stopWords:
            continue
        if w in frequencyTable:
            frequencyTable[w] += 1
        else:
            frequencyTable[w] = 1
    
    return frequencyTable

    

"""scoring sentences by its words
    Basic algorithm: adding the frequency of every non-stop word in a sentence divided by total no of words in a sentence."""
def scoring_input_sentences(input_sentences, freqTable) -> dict:
    

    sentenceScoreDict = dict()

    for s in input_sentences:
        word_count_in_sentence = (len(word_tokenize(s)))
        word_count_in_sentence_except_stop_words = 0
        for wordValue in freqTable:
            if wordValue in s.lower():
                word_count_in_sentence_except_stop_words += 1
                if s[:10] in sentenceScoreDict:
                    sentenceScoreDict[s[:10]] += freqTable[wordValue]
                else:
                    sentenceScoreDict[s[:10]] = freqTable[wordValue]

        sentenceScoreDict[s[:10]] = sentenceScoreDict[s[:10]] / word_count_in_sentence_except_stop_words
        
    return sentenceScoreDict
    

"""To calculate the average score"""
def calculate_average_score(sentenceScore) -> int:
    
    sumScores = 0
    for s in sentenceScore:
        sumScores += sentenceScore[s]

    avg_score = (sumScores / len(sentenceScore))
    

"""To generate summary output"""
def create_summary(sentences, sentenceScore, threshold):
    sentence_counter = 0
    summary_output = ''

    for s in sentences:
        if s[:10] in sentenceScore and sentenceScore[s[:10]] > (threshold):
            summary_output += " " + s
            sentence_counter += 1

    return summary_output


def summarization(input_text):
    
    freq_table = generate_frequency_table(input_text)

    sentences = sent_tokenize(input_text)

    sentence_scores = scoring_input_sentences(sentences, freq_table)

    threshold_value = calculate_average_score(sentence_scores)

    summary_output = create_summary(sentences, sentence_scores, 0.99*threshold_value)

    return summary_output


def read_and_write():
    
    path = "C:\\Users\\gsree\\OneDrive\\Desktop\\Sentiment Analyis Output.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    m_row = sheet_obj.max_row 
    for i in range(2, m_row+1): 
        cell_obj = sheet_obj.cell(row = i, column = 1) 
        result = summarization(cell_obj.value)
        ci = sheet_obj.cell(row = i, column = 8)
        ci.value = result
        print(result)
        wb_obj.save("C:\\Users\\gsree\\OneDrive\\Desktop\\Sentiment Analyis Output.xlsx") 
    
if __name__ == '__main__':
    
    read_and_write()